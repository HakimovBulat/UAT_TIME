import logging
from telegram import ReplyKeyboardMarkup, InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import Application, MessageHandler, filters, CommandHandler, ConversationHandler, CallbackQueryHandler
from timetable import send_day_timetable, send_ring_time, CURRENT_WEEK_NUMBER
from openpyxl import load_workbook
import datetime


logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)


logger = logging.getLogger(__name__)
reply_keyboard = [["/today", "/tomorrow"], ["/ring", "/select_day"]]
markup = ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=False)
TOKEN = "7521976097:AAHy6d-dRYbM0xB4KkNT5fiTQ7juhUe0NGI"
GROUP = "ИСП(п)3122"
WEEK_NAMES =  ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
WEEK_NUMBER = 0
GROUP = "ИСП(п)3122"


async def start(update, context):
    user = update.effective_user
    faculties = load_workbook("1 семестр Расписание 3 курса.xlsx").sheetnames
    keyboard = []
    for faculty in faculties:
        keyboard.append([InlineKeyboardButton(faculty, callback_data=faculty)])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_html(
        rf"Привет, {user.mention_html()}! Выбери специальность: ", reply_markup=reply_markup
    )
    return 2

select_faculty
async def select_group(update, context):
    query = update.callback_query
    await query.answer()
    faculty = query.data
    keyboard = []
    for group in load_workbook("1 семестр Расписание 3 курса.xlsx")[faculty]["1"]:
        if group.value is not None and group.value not in ["День недели", "Время", "№ пары"]:
            keyboard.append([InlineKeyboardButton(group.value, callback_data=group.value)])
    reply_markup = InlineKeyboardMarkup(keyboard)
    # await update.message.reply_text(f"Выберите группу: ", reply_markup=reply_markup)
    await query.edit_message_text(text=f"Выбери группу: ", reply_markup=reply_markup)
    return ConversationHandler.END


async def button_group(update, context) -> None:
    global GROUP
    query = update.callback_query
    await query.answer()
    GROUP = query.data
    await query.edit_message_text(text="OK")


async def today(update, context):
    day = WEEK_NAMES[datetime.datetime.today().weekday()]
    if datetime.datetime.today().weekday() != 6:
        lessons_str = '\n'.join(send_day_timetable('ИСП(п)3122', day))
        day = day[:-1] + "у" if day[-1] == 'а' else day
        message = f"Расписание на {day}: \n{lessons_str}"
    else:
        message = f"Расписание на {day}: \nПар нет, так что можно отдохнуть"
    await update.message.reply_text(message)


async def select_week(update, context):
    await update.message.reply_text(f"Напишите номер нужной недели для просмотра расписания (текущая - {CURRENT_WEEK_NUMBER})")
    return 1


async def select_day(update, context):
    global WEEK_NUMBER
    WEEK_NUMBER = int(update.message.text)
    keyboard = [
        [
            InlineKeyboardButton("Понедельник", callback_data="Понедельник"),
            InlineKeyboardButton("Вторник", callback_data="Вторник"),
        ],
        [
            InlineKeyboardButton("Среда", callback_data="Среда"),
            InlineKeyboardButton("Четверг", callback_data="Четверг"),
        ],
        [
            InlineKeyboardButton("Пятница", callback_data="Пятница"),
            InlineKeyboardButton("Суббота", callback_data="Суббота"),
        ],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите день:", reply_markup=reply_markup)
    return ConversationHandler.END


async def stop(update, context):
    return ConversationHandler.END


async def button_day(update, context) -> None:
    global WEEK_NUMBER
    query = update.callback_query
    await query.answer()
    day = query.data

    lessons_str = '\n'.join(send_day_timetable('ИСП(п)3122', day, WEEK_NUMBER))
    day = day[:-1] + "у" if day[-1] == 'а' else day
    message = f"Расписание на {day}: \n{lessons_str}"

    await query.edit_message_text(text=message)


async def tomorrow(update, context):
    day = WEEK_NAMES[(datetime.datetime.today().weekday() + 1) % 7]
    if datetime.datetime.today().weekday() != 5:
        lessons_str = '\n'.join(send_day_timetable('ИСП(п)3122', day, ))
        day = day[:-1] + "у" if day[-1] == 'а' else day
        message = f"Расписание на {day}: \n{lessons_str}"
    else:
        message = f"Расписание на {day}: \nПар нет, так что можно отдохнуть"
    await update.message.reply_text(message)


async def echo(update, context):
    await update.message.reply_text(
        "Я не понимаю, используйте кнопки меню или /help"
    )


async def ring(update, context):
    message = f"Звонок в {send_ring_time()}"
    if message == "Звонок в Пары закончились":
        message = "Пары закончились"
    await update.message.reply_text(message)


async def help_command(update, context):
    await update.message.reply_text(
        "Я - бот для просмотра расписаний УАТ. Вот мои команды:  \
            \n/help - Список команд \
            \n/select - выбрать конкретныю неделю и день\
            \n/today - Расписание на сегодня \
            \n/tomorrow - Расписание на завтра \
            \n/ring - Ближайший звонок",
            reply_markup=markup,
    )


def main():
    application = Application.builder().token(TOKEN).build()
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("faculty", select_faculty))
    application.add_handler(CallbackQueryHandler(button_faculty))
    application.add_handler(CommandHandler("group", select_group))
    application.add_handler(CallbackQueryHandler(button_group))

    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("ring", ring))
    application.add_handler(CommandHandler("today", today))
    application.add_handler(CommandHandler("tomorrow", tomorrow))
    select_day_handler = ConversationHandler(
        entry_points=[CommandHandler("select_day", select_week)],
        states={
            1: [MessageHandler(filters.TEXT & ~filters.COMMAND, select_day)],
        },
        fallbacks=[CommandHandler("stop", stop)],
    )
    application.add_handler(CallbackQueryHandler(button_day))
    application.add_handler(select_day_handler)



    text_handler = MessageHandler(filters.TEXT & ~filters.COMMAND, echo)
    application.add_handler(text_handler)
    

    application.run_polling()


if __name__ == "__main__":
    main()