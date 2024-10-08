import requests
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
import datetime

url = 'https://docs.google.com/spreadsheets/d/1bckdpp4i-J0iFszaE-tqODnVhfNutyHKCW5wdFFD8-Y/export?format=xlsx'
response = requests.get(url)
with open('1 семестр Расписание 3 курса.xlsx', "wb") as file:
	file.write(response.content)

WEEK_NAMES =  ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
GROUPS = ["ИСП(п)3122", "ИСП(с)3222", "ИСП(п)3322", "ИСП(с)3422", "ИСП(с)3522", "ИСП(с)3622"]
CURRENT_WEEK_NUMBER = 2

FACULTY = "ИСП ПР"


def send_day_timetable(group : str, week_day : str, week_number : int=CURRENT_WEEK_NUMBER) -> list:
    if week_number is None:
        week_number = CURRENT_WEEK_NUMBER
    week_days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
    wb = load_workbook("1 семестр Расписание 3 курса.xlsx")
    ws = wb[FACULTY]
    print(week_number)
    row_index = week_days.index(week_day)
    group_index = GROUPS.index(group)
    count_group_week = 0
    for cell in ws["2"]:
        if cell.value == week_number and count_group_week == group_index:
            col_index = cell.column
            break
        elif cell.value == week_number and count_group_week != group_index:
            count_group_week += 1
    lessons = []
    for row in ws.iter_rows(min_row=row_index * 6 + 3, max_col=col_index, max_row=row_index * 6 + 8, min_col=col_index):
        for cell in row:
            if cell.value:
                lessons.append(cell.value)
            else:
                lessons.append("------------------------")
    if lessons.count("------------------------") == 6:
        return ["Пар нет, так что можно отдохнуть"]
    for i in range(len(lessons)):
        lessons[i] = f"{i+1}) {lessons[i]}"
    return lessons


def send_ring_time():
    now = datetime.datetime.now()
    if now <= now.replace(hour=8, minute=0):
        return "8:00"
    elif now <= now.replace(hour=9, minute=30):
        return "9:30"
    elif now <= now.replace(hour=9, minute=40):
        return "9:40"
    elif now <= now.replace(hour=11, minute=10):
        return "11:10"
    elif now <= now.replace(hour=12, minute=0):
        return "12:00"
    elif now <= now.replace(hour=13, minute=30):
        return "13:30"
    elif now <= now.replace(hour=13, minute=40):
        return "13:40"
    elif now <= now.replace(hour=15, minute=10):
        return "15:10"
    elif now <= now.replace(hour=15, minute=50):
        return "15:50"
    elif now <= now.replace(hour=17, minute=20):
        return "17:20"
    elif now <= now.replace(hour=17, minute=30):
        return "17:30"
    elif now <= now.replace(hour=19, minute=0):
        return "19:00"
    return "Пары закончились"